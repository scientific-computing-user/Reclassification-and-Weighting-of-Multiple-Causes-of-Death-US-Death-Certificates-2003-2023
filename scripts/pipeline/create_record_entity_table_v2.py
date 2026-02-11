#!/usr/bin/env python3
"""
Create Record Axis vs Entity Axis Comparison Tables

This script analyzes death certificate data and produces two tables:

TABLE 1: Comparison Table
Compares causes of death classified in broad disease categories between:
  - Record Axis (rUDS): NCHS/ACME automated system's final determination
  - Entity Axis (eUDS): Physician's original coding on the death certificate

Columns for each axis:
  - Underlying: First position (underlying cause of death)
  - Contributing: Not in first position (contributing causes)
  - Only: Single-cause deaths (string length = 1)
  - Any: Appears anywhere on the certificate

TABLE 2: Transition Matrix
Shows how underlying causes transition from Entity Axis to Record Axis:
  - Rows = Record Axis underlying cause (what ACME determined)
  - Columns = Entity Axis underlying cause (what physician certified)
  - Cell values = Count of deaths with that Entity->Record combination

Input CSV format:
    Count,rUDS,eUDS,year,age_group
    169037,C,C,2003,GE65
    ...

Usage:
    python create_record_entity_table.py <input.csv> --years 2003-2023
    python create_record_entity_table.py <input.csv>  # Uses all years

Author: Claude & Michael Levitt
Date: January 2025
"""

import pandas as pd
import numpy as np
from collections import defaultdict
import sys
import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION
# ============================================================================

ALLOWED_LETTERS = ["B", "C", "N", "R", "E", "D", "V", "P", "T", "S", "A", "X", "F", "H", "U"]
ALLOWED_SET = set(ALLOWED_LETTERS)

DISEASE_NAMES = {
    'B': 'Circulatory',
    'C': 'Cancer',
    'N': 'Other Natural',
    'R': 'Respiratory',
    'E': 'Endocrine*',
    'D': 'Digestive',
    'V': 'COVID-19',
    'P': 'Drug Poisoning',
    'T': 'Transport',
    'S': 'Suicide',
    'A': 'Alcohol-related',
    'X': 'Other External',
    'F': 'Falls',
    'H': 'Homicide',
    'U': 'Uncertain'
}

# Order for output
DISEASE_ORDER = ['B', 'C', 'N', 'R', 'E', 'D', 'V', 'P', 'T', 'S', 'A', 'X', 'F', 'H', 'U']

AGE_GROUPS = ["ALL", "LT65", "GE65"]

# ============================================================================
# ARGUMENT PARSING
# ============================================================================

def parse_arguments():
    parser = argparse.ArgumentParser(
        description='Create Record Axis vs Entity Axis comparison table.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python create_record_entity_table.py data.csv --years 2003-2023
    python create_record_entity_table.py data.csv  # Uses all years
    python create_record_entity_table.py data.csv --label "USA 2003-2023"
        """
    )
    
    parser.add_argument('filename', help='Input CSV file with Count,rUDS,eUDS,age_group (year column optional)')
    
    parser.add_argument('--years', '-y', type=str, default=None,
                        help='Year range to filter (e.g., 2003-2023). Only used if year column exists.')
    
    parser.add_argument('--label', '-l', type=str, default=None,
                        help='Label for table title (e.g., "USA 2003-2023")')
    
    parser.add_argument('--output', '-o', type=str, default=None,
                        help='Output filename (default: Record_Entity_Comparison.xlsx)')
    
    return parser.parse_args()


def parse_year_range(year_str):
    """Parse year range string into start and end years."""
    if year_str is None:
        return None, None
    parts = year_str.split('-')
    if len(parts) != 2:
        print(f"ERROR: Invalid year range: {year_str}")
        sys.exit(1)
    return int(parts[0]), int(parts[1])


# ============================================================================
# ANALYSIS FUNCTIONS
# ============================================================================

def analyze_uds_column(df, uds_column, count_column='Count'):
    """
    Analyze a UDS column to get Underlying, Contributing, Only, and Any counts.
    
    Returns dict with letter stats:
    - Underlying (first): count where letter is first
    - Contributing (not_first): count where letter appears but not first  
    - Only: count where letter is only letter (Ls=1)
    - Any (anywhere): count where letter appears anywhere
    """
    letter_stats = {letter: {
        'Underlying': 0,
        'Contributing': 0,
        'Only': 0,
        'Any': 0
    } for letter in ALLOWED_LETTERS}
    
    for _, row in df.iterrows():
        uds = row[uds_column]
        count = row[count_column]
        
        if not isinstance(uds, str) or len(uds) == 0:
            continue
        
        Ls = len(uds)
        
        # Track unique letters seen in this record
        seen = set()
        
        for i, letter in enumerate(uds):
            if letter not in ALLOWED_SET:
                continue
            
            # Only count each letter once per record for these metrics
            if letter not in seen:
                if i == 0:
                    letter_stats[letter]['Underlying'] += count
                else:
                    letter_stats[letter]['Contributing'] += count
                
                if Ls == 1:
                    letter_stats[letter]['Only'] += count
                
                letter_stats[letter]['Any'] += count
                seen.add(letter)
    
    return letter_stats


def analyze_transition_matrix(df, count_column='Count'):
    """
    Create transition matrix from Entity Axis (columns) to Record Axis (rows).
    
    For each death, looks at:
    - Entity underlying cause (first letter of eUDS)
    - Record underlying cause (first letter of rUDS)
    
    Returns 2D dict: matrix[record_letter][entity_letter] = count
    """
    matrix = {r_letter: {e_letter: 0 for e_letter in ALLOWED_LETTERS} 
              for r_letter in ALLOWED_LETTERS}
    
    for _, row in df.iterrows():
        eUDS = row['eUDS']
        rUDS = row['rUDS']
        count = row[count_column]
        
        if not isinstance(eUDS, str) or len(eUDS) == 0:
            continue
        if not isinstance(rUDS, str) or len(rUDS) == 0:
            continue
        
        e_underlying = eUDS[0]
        r_underlying = rUDS[0]
        
        if e_underlying in ALLOWED_SET and r_underlying in ALLOWED_SET:
            matrix[r_underlying][e_underlying] += count
    
    return matrix


def run_analysis(df, start_year=None, end_year=None):
    """
    Run analysis on data for a specific year range.
    
    Returns dict with results for each age group.
    """
    # Filter by year if year column exists and year range is specified
    if start_year is not None and 'year' in df.columns:
        df_filtered = df[(df['year'] >= start_year) & (df['year'] <= end_year)].copy()
    else:
        df_filtered = df.copy()
    
    total_deaths = df_filtered['Count'].sum()
    print(f"  Records: {len(df_filtered):,}, Deaths: {total_deaths:,}")
    
    results = {}
    
    # Check if age_group column exists (handle different naming)
    age_col = None
    for col in ['age_group', 'age-group', 'AgeGroup', 'agegroup']:
        if col in df_filtered.columns:
            age_col = col
            break
    
    for age_group in AGE_GROUPS:
        if age_group == 'ALL':
            df_subset = df_filtered
        else:
            if age_col is None:
                print(f"    [SKIP] {age_group}: No age_group column found")
                continue
            df_subset = df_filtered[df_filtered[age_col] == age_group]
        
        if len(df_subset) == 0:
            print(f"    [SKIP] {age_group}: No data")
            continue
        
        subset_deaths = df_subset['Count'].sum()
        
        # Analyze rUDS (Record Axis)
        rUDS_stats = analyze_uds_column(df_subset, 'rUDS')
        
        # Analyze eUDS (Entity Axis)
        eUDS_stats = analyze_uds_column(df_subset, 'eUDS')
        
        # Analyze transition matrix (Entity -> Record)
        transition_matrix = analyze_transition_matrix(df_subset)
        
        results[age_group] = {
            'rUDS': rUDS_stats,
            'eUDS': eUDS_stats,
            'transition': transition_matrix,
            'deaths': subset_deaths
        }
        
        print(f"    {age_group}: {subset_deaths:,} deaths")
    
    return results


# ============================================================================
# EXCEL OUTPUT FUNCTIONS
# ============================================================================

# Styles
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def format_number(val):
    """Format number with comma separators."""
    return f"{int(val):,}"


def create_comparison_table(results, year_range_str, output_filename):
    """
    Create the Record Axis vs Entity Axis comparison table.
    
    Format matches the image:
    - Record Axis: Underlying, Contributing, Only, Any
    - Entity Axis: Underlying, Contributing, Only, Any
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Record_vs_Entity"
    
    row_num = 1
    
    for age_group in ['ALL', 'LT65', 'GE65']:
        if age_group not in results:
            continue
        
        rUDS_stats = results[age_group]['rUDS']
        eUDS_stats = results[age_group]['eUDS']
        total_deaths = results[age_group]['deaths']
        
        # Age group label
        age_label = {
            'ALL': 'All Ages',
            'LT65': 'Under 65 years',
            'GE65': '65 years and older'
        }[age_group]
        
        # Title
        title = f"Table 1. Causes of death classified in broad disease categories in Record Axis and Entity Axis, USA {year_range_str}"
        if age_group != 'ALL':
            title = f"Table 1. Causes of death classified in broad disease categories in Record Axis and Entity Axis, USA {year_range_str} ({age_label})"
        
        ws.cell(row=row_num, column=1, value=title)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
        row_num += 2
        
        # Sub-headers: Record Axis | Entity Axis
        ws.cell(row=row_num, column=2, value="Record Axis")
        ws.cell(row=row_num, column=2).font = HEADER_FONT
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        
        ws.cell(row=row_num, column=6, value="Entity Axis")
        ws.cell(row=row_num, column=6).font = HEADER_FONT
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=row_num, start_column=6, end_row=row_num, end_column=9)
        row_num += 1
        
        # Column headers
        headers = ['Disease', 'Underlying', 'Contributing', 'Only', 'Any', 
                   'Underlying', 'Contributing', 'Only', 'Any']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        row_num += 1
        
        # Data rows
        for letter in DISEASE_ORDER:
            disease_name = DISEASE_NAMES.get(letter, letter)
            
            r_stats = rUDS_stats[letter]
            e_stats = eUDS_stats[letter]
            
            row_data = [
                disease_name,
                format_number(r_stats['Underlying']),
                format_number(r_stats['Contributing']),
                format_number(r_stats['Only']),
                format_number(r_stats['Any']),
                format_number(e_stats['Underlying']),
                format_number(e_stats['Contributing']),
                format_number(e_stats['Only']),
                format_number(e_stats['Any'])
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = THIN_BORDER
                if col == 1:
                    cell.font = Font(bold=True)
                else:
                    cell.alignment = Alignment(horizontal='right')
            
            row_num += 1
        
        # Total row
        total_r_underlying = sum(rUDS_stats[l]['Underlying'] for l in DISEASE_ORDER)
        total_r_contributing = sum(rUDS_stats[l]['Contributing'] for l in DISEASE_ORDER)
        total_r_only = sum(rUDS_stats[l]['Only'] for l in DISEASE_ORDER)
        total_r_any = sum(rUDS_stats[l]['Any'] for l in DISEASE_ORDER)
        
        total_e_underlying = sum(eUDS_stats[l]['Underlying'] for l in DISEASE_ORDER)
        total_e_contributing = sum(eUDS_stats[l]['Contributing'] for l in DISEASE_ORDER)
        total_e_only = sum(eUDS_stats[l]['Only'] for l in DISEASE_ORDER)
        total_e_any = sum(eUDS_stats[l]['Any'] for l in DISEASE_ORDER)
        
        total_data = [
            'Total',
            format_number(total_r_underlying),
            format_number(total_r_contributing),
            format_number(total_r_only),
            format_number(total_r_any),
            format_number(total_e_underlying),
            format_number(total_e_contributing),
            format_number(total_e_only),
            format_number(total_e_any)
        ]
        
        for col, val in enumerate(total_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(bold=True)
            if col > 1:
                cell.alignment = Alignment(horizontal='right')
        
        row_num += 2
        
        # Footnote
        ws.cell(row=row_num, column=1, value="*endocrine nutritional and metabolic")
        ws.cell(row=row_num, column=1).font = Font(italic=True)
        row_num += 3  # Extra space between age groups
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 16
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 14
    
    wb.save(output_filename)
    print(f"  Saved: {output_filename}")


def create_transition_matrix_table(results, year_range_str, output_filename):
    """
    Create the transition matrix table showing Entity Axis (columns) to Record Axis (rows).
    
    Each cell shows the count of deaths where:
    - Entity Axis had column disease as underlying cause
    - Record Axis had row disease as underlying cause
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transition_Matrix"
    
    # Short names for column headers (to fit)
    SHORT_NAMES = {
        'B': 'Circulatory',
        'C': 'Cancer',
        'N': 'Other Natural',
        'R': 'Respiratory',
        'E': 'Endocrine*',
        'D': 'Digestive',
        'V': 'COVID-19',
        'P': 'Drug Poisoning',
        'T': 'Transport',
        'S': 'Suicide',
        'A': 'Alcohol',
        'X': 'Other External',
        'F': 'Falls',
        'H': 'Homicide',
        'U': 'Uncertain'
    }
    
    row_num = 1
    
    for age_group in ['ALL', 'LT65', 'GE65']:
        if age_group not in results:
            continue
        
        transition = results[age_group]['transition']
        
        age_label = {
            'ALL': 'All Ages',
            'LT65': 'Under 65 years',
            'GE65': '65 years and older'
        }[age_group]
        
        # Title
        if age_group == 'ALL':
            title = f"Table 2. Transition matrix from Entity Axis (columns) to Record Axis (rows), USA {year_range_str}"
        else:
            title = f"Table 2. Transition matrix from Entity Axis (columns) to Record Axis (rows), USA {year_range_str} ({age_label})"
        
        ws.cell(row=row_num, column=1, value=title)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(DISEASE_ORDER)+1)
        row_num += 2
        
        # Header row with ENTITY on top-left corner indicator
        ws.cell(row=row_num, column=1, value="ENTITY")
        ws.cell(row=row_num, column=1).font = HEADER_FONT
        ws.cell(row=row_num+1, column=1, value="RECORD")
        ws.cell(row=row_num+1, column=1).font = HEADER_FONT
        
        # Column headers (Entity Axis causes)
        for col_idx, letter in enumerate(DISEASE_ORDER, 2):
            cell = ws.cell(row=row_num, column=col_idx, value=SHORT_NAMES[letter])
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        row_num += 1
        
        # Data rows (Record Axis causes)
        for r_letter in DISEASE_ORDER:
            # Row label
            cell = ws.cell(row=row_num, column=1, value=DISEASE_NAMES[r_letter])
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            
            # Data cells
            for col_idx, e_letter in enumerate(DISEASE_ORDER, 2):
                val = transition[r_letter][e_letter]
                cell = ws.cell(row=row_num, column=col_idx, value=format_number(val))
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='right')
            
            row_num += 1
        
        row_num += 1
        
        # Footnote
        ws.cell(row=row_num, column=1, value="*endocrine nutritional and metabolic")
        ws.cell(row=row_num, column=1).font = Font(italic=True)
        row_num += 3
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 14
    for col_idx in range(2, len(DISEASE_ORDER)+2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    
    wb.save(output_filename)
    print(f"  Saved: {output_filename}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    args = parse_arguments()
    
    print("=" * 70)
    print("RECORD AXIS vs ENTITY AXIS COMPARISON TABLE")
    print("=" * 70)
    
    # Parse year range
    start_year, end_year = parse_year_range(args.years)
    
    # Determine label for table title
    if args.label:
        year_range_str = args.label
    elif start_year:
        year_range_str = f"{start_year}-{end_year}"
    else:
        year_range_str = "All Years"
    
    # Read data
    print(f"\nReading: {args.filename}")
    df = pd.read_csv(args.filename, keep_default_na=False, na_values=[''])
    print(f"  Loaded {len(df):,} records")
    print(f"  Columns: {list(df.columns)}")
    
    # Check required columns
    required = ['Count', 'rUDS', 'eUDS']
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"ERROR: Missing required columns: {missing}")
        print(f"  Available columns: {list(df.columns)}")
        sys.exit(1)
    
    total_deaths = df['Count'].sum()
    print(f"  Total deaths in file: {total_deaths:,}")
    
    # Run analysis
    print(f"\nAnalyzing...")
    results = run_analysis(df, start_year, end_year)
    
    # Output filename
    if args.output:
        output_file = args.output
    else:
        if start_year:
            output_file = f"Record_Entity_Comparison_{start_year}-{end_year}.xlsx"
        else:
            output_file = "Record_Entity_Comparison.xlsx"
    
    # Transition matrix output filename
    transition_file = output_file.replace('.xlsx', '_Transition.xlsx')
    
    # Create tables
    print("\n" + "=" * 70)
    print("CREATING EXCEL TABLES")
    print("=" * 70)
    
    create_comparison_table(results, year_range_str, output_file)
    create_transition_matrix_table(results, year_range_str, transition_file)
    
    print("\n" + "=" * 70)
    print("DONE")
    print("=" * 70)
    print(f"\nOutput files:")
    print(f"  Table 1 (Comparison): {output_file}")
    print(f"  Table 2 (Transition): {transition_file}")


if __name__ == "__main__":
    main()
