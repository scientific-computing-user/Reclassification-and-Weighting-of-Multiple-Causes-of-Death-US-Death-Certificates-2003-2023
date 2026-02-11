#!/usr/bin/env python3
"""
Create Disease Weight Tables from rUDS/rADS analysis.

This script analyzes death certificate data and produces two Excel tables:
  Table 1: ALL, LT65, GE65 stacked vertically (single time period)
  Table 2: LT65 and GE65 with All Years and Pandemic Years side by side

Weighting schemes:
  W0:  Underlying cause only (no weighting)
  W1:  50% UCoD + 50% split among contributing causes (unique letters)
  W2:  Equal weight among unique disease letters
  W2A: Equal weight among ALL ICD-level causes (including repeats)

Input CSV format:
    Count,rUDS,rADS,year,age_group
    169037,C,C,2003,GE65
    ...

Usage:
    python create_disease_weight_tables.py <input.csv> [--all-years 2003-2023] [--pandemic-years 2020-2023]

Author: Claude & Michael Levitt
Date: January 2025
"""

import pandas as pd
import numpy as np
from collections import defaultdict
import sys
import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ============================================================================
# CONFIGURATION
# ============================================================================

ALLOWED_LETTERS = ["B", "C", "N", "R", "E", "D", "V", "P", "T", "S", "A", "X", "F", "H", "U"]
ALLOWED_SET = set(ALLOWED_LETTERS)

DISEASE_NAMES = {
    'B': 'Circulatory',
    'C': 'Cancer',
    'N': 'Other Natural',
    'R': 'Respiratory',
    'E': 'Endocrine*',
    'D': 'Digestive',
    'V': 'COVID-19',
    'P': 'Drug Poisoning',
    'T': 'Transport',
    'S': 'Suicide',
    'A': 'Alcohol-related',
    'X': 'Other External',
    'F': 'Falls',
    'H': 'Homicide',
    'U': 'Unknown'
}

# Order for output (excludes U which is rare)
DISEASE_ORDER = ['B', 'C', 'N', 'R', 'E', 'D', 'V', 'P', 'T', 'S', 'A', 'X', 'F', 'H']

AGE_GROUPS = ["ALL", "LT65", "GE65"]

# ============================================================================
# ARGUMENT PARSING
# ============================================================================

def parse_arguments():
    parser = argparse.ArgumentParser(
        description='Create disease weight tables from death certificate data.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python create_disease_weight_tables.py data.csv
    python create_disease_weight_tables.py data.csv --all-years 2003-2023 --pandemic-years 2020-2023
    python create_disease_weight_tables.py data.csv --output my_tables
        """
    )
    
    parser.add_argument('filename', help='Input CSV file with Count,rUDS,rADS,year,age_group')
    
    parser.add_argument('--all-years', type=str, default='2003-2023',
                        help='Year range for "All Years" (default: 2003-2023)')
    
    parser.add_argument('--pandemic-years', type=str, default='2020-2023',
                        help='Year range for "Pandemic Years" (default: 2020-2023)')
    
    parser.add_argument('--output', '-o', type=str, default=None,
                        help='Output filename prefix (default: Disease_Weights_Tables)')
    
    return parser.parse_args()


def parse_year_range(year_str):
    """Parse year range string into start and end years."""
    if year_str is None:
        return None, None
    parts = year_str.split('-')
    if len(parts) != 2:
        print(f"ERROR: Invalid year range: {year_str}")
        sys.exit(1)
    return int(parts[0]), int(parts[1])


# ============================================================================
# ANALYSIS FUNCTIONS
# ============================================================================

def analyze_ads_column(df, ads_column, count_column='Count'):
    """
    Analyze an ADS column (with or without repeats).
    
    W1 weighting: first letter gets 0.5, others share 0.5 equally
    W2A weighting: each letter gets 1/Ls for each occurrence
    
    Returns dict with letter stats.
    """
    letter_stats = {letter: {
        'First': 0,
        'Not_First': 0,
        'Only': 0,
        'W1': 0.0,
        'W2A': 0.0,
        'Anywhere': 0,
        'Ls_sum': 0,
        'Ls_count': 0
    } for letter in ALLOWED_LETTERS}
    
    for _, row in df.iterrows():
        ads = row[ads_column]
        count = row[count_column]
        
        if not isinstance(ads, str) or len(ads) == 0:
            continue
        
        # Fix: If first letter is duplicated in second position, remove first letter
        # e.g., "CCBN" -> "CBN", "BB" -> "B"
        if len(ads) >= 2 and ads[0] == ads[1]:
            ads = ads[1:]
        
        Ls = len(ads)
        first_letter = ads[0]
        
        # Track Ls distribution for first letter
        if first_letter in ALLOWED_SET:
            letter_stats[first_letter]['Ls_sum'] += Ls * count
            letter_stats[first_letter]['Ls_count'] += count
        
        # W2A weights: each position gets count/Ls
        w2_per_position = count / Ls
        
        # W1 weights: first position gets 0.5, others share 0.5 equally
        if Ls == 1:
            w1_per_first = count
            w1_per_other = 0
        else:
            w1_per_first = count * 0.5
            w1_per_other = count * 0.5 / (Ls - 1)
        
        # Process each position
        seen = set()
        for i, letter in enumerate(ads):
            if letter not in ALLOWED_SET:
                continue
            
            # W1: each position gets its weight
            if i == 0:
                letter_stats[letter]['W1'] += w1_per_first
            else:
                letter_stats[letter]['W1'] += w1_per_other
            
            # W2A: each occurrence gets w2_per_position
            letter_stats[letter]['W2A'] += w2_per_position
            
            # Anywhere: count ALL occurrences
            letter_stats[letter]['Anywhere'] += count
            
            # First/Not_First/Only: only count once per letter per record
            if letter not in seen:
                if i == 0:
                    letter_stats[letter]['First'] += count
                else:
                    letter_stats[letter]['Not_First'] += count
                
                if Ls == 1:
                    letter_stats[letter]['Only'] += count
                
                seen.add(letter)
    
    # Compute Mean_Ls
    for letter in ALLOWED_LETTERS:
        if letter_stats[letter]['Ls_count'] > 0:
            letter_stats[letter]['Mean_Ls'] = letter_stats[letter]['Ls_sum'] / letter_stats[letter]['Ls_count']
        else:
            letter_stats[letter]['Mean_Ls'] = 0.0
    
    return letter_stats


def run_analysis(df, start_year, end_year):
    """
    Run analysis on data for a specific year range.
    
    Returns dict with results for each age group, containing rUDS and rADS stats.
    """
    # Filter by year
    if start_year is not None:
        df_filtered = df[(df['year'] >= start_year) & (df['year'] <= end_year)].copy()
    else:
        df_filtered = df.copy()
    
    total_deaths = df_filtered['Count'].sum()
    print(f"  Year range {start_year}-{end_year}: {len(df_filtered):,} records, {total_deaths:,} deaths")
    
    results = {}
    
    for age_group in AGE_GROUPS:
        if age_group == 'ALL':
            df_subset = df_filtered
        else:
            df_subset = df_filtered[df_filtered['age_group'] == age_group]
        
        if len(df_subset) == 0:
            print(f"    [SKIP] {age_group}: No data")
            continue
        
        subset_deaths = df_subset['Count'].sum()
        
        # Analyze rUDS (unique letters)
        rUDS_stats = analyze_ads_column(df_subset, 'rUDS')
        
        # Analyze rADS (all letters with repeats)
        rADS_stats = analyze_ads_column(df_subset, 'rADS')
        
        results[age_group] = {
            'rUDS': rUDS_stats,
            'rADS': rADS_stats,
            'deaths': subset_deaths
        }
        
        print(f"    {age_group}: {subset_deaths:,} deaths")
    
    return results


# ============================================================================
# EXCEL OUTPUT FUNCTIONS
# ============================================================================

# Styles
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def format_value(val, w0_val):
    """Format value in thousands with % change from W0 in parentheses."""
    val = float(val)
    w0 = float(w0_val)
    val_k = val / 1000
    pct = (val / w0 * 100) if w0 > 0 else 0
    delta = pct - 100
    if delta >= 0:
        return f"{val_k:,.0f} (+{delta:.0f})"
    else:
        return f"{val_k:,.0f} ({delta:.0f})"


def format_w0(val):
    """Format W0 value in thousands."""
    return f"{float(val)/1000:,.0f}"


def create_table1(results_all, results_pandemic, all_years_str, pandemic_years_str, output_filename):
    """
    Create Table 1: ALL ages with All Years and Pandemic Years side by side.
    
    Columns: Disease, W0, W1, W2, W2A (All Years), W0, W1, W2, W2A (Pandemic Years)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Disease_Weights_All_Ages"
    
    row_num = 1
    
    # Title
    ws.cell(row=row_num, column=1, 
            value=f"Table 4. Death counts for broad disease categories for {all_years_str} and for the pandemic years {pandemic_years_str} with different weighting schemes")
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
    row_num += 1
    
    ws.cell(row=row_num, column=1, 
            value="(W0: no weight, only UCoD counted; W1: 50% weight for UCoD and 50% shared equally among other causes; W2: weight shared equally among all causes; W2A: weight shared equally among all causes but considering ICD-level causes rather than just broad categories)")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
    row_num += 2
    
    # Section header
    ws.cell(row=row_num, column=1, value="Counts for All Ages, in thousands (% of W0)")
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
    row_num += 1
    
    # Sub-headers: All Years | Pandemic Years
    ws.cell(row=row_num, column=2, value=f"All Years ({all_years_str})")
    ws.cell(row=row_num, column=2).font = HEADER_FONT
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
    
    ws.cell(row=row_num, column=6, value=f"Pandemic Years ({pandemic_years_str})")
    ws.cell(row=row_num, column=6).font = HEADER_FONT
    ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row_num, start_column=6, end_row=row_num, end_column=9)
    row_num += 1
    
    # Column headers
    headers = ['Disease', 'W0', 'W1', 'W2', 'W2A', 'W0', 'W1', 'W2', 'W2A']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    row_num += 1
    
    # Get ALL ages data
    rUDS_all = results_all['ALL']['rUDS']
    rADS_all = results_all['ALL']['rADS']
    rUDS_pan = results_pandemic['ALL']['rUDS'] if 'ALL' in results_pandemic else None
    rADS_pan = results_pandemic['ALL']['rADS'] if 'ALL' in results_pandemic else None
    
    # Data rows
    for letter in DISEASE_ORDER:
        # All years data
        w0_all = rUDS_all[letter]['First']
        w1_all = rUDS_all[letter]['W1']
        w2_all = rUDS_all[letter]['W2A']
        w2a_all = rADS_all[letter]['W2A']
        
        # Pandemic years data
        if rUDS_pan and rADS_pan:
            w0_pan = rUDS_pan[letter]['First']
            w1_pan = rUDS_pan[letter]['W1']
            w2_pan = rUDS_pan[letter]['W2A']
            w2a_pan = rADS_pan[letter]['W2A']
        else:
            w0_pan = w1_pan = w2_pan = w2a_pan = 0
        
        disease_name = DISEASE_NAMES.get(letter, letter)
        
        row_data = [
            disease_name,
            format_w0(w0_all),
            format_value(w1_all, w0_all),
            format_value(w2_all, w0_all),
            format_value(w2a_all, w0_all),
            format_w0(w0_pan),
            format_value(w1_pan, w0_pan) if w0_pan > 0 else "0 (+0)",
            format_value(w2_pan, w0_pan) if w0_pan > 0 else "0 (+0)",
            format_value(w2a_pan, w0_pan) if w0_pan > 0 else "0 (+0)"
        ]
        
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = THIN_BORDER
            if col == 1:
                cell.font = Font(bold=True)
            else:
                cell.alignment = Alignment(horizontal='right')
        
        row_num += 1
    
    # Footnote
    row_num += 1
    ws.cell(row=row_num, column=1, value="*endocrine nutritional and metabolic")
    ws.cell(row=row_num, column=1).font = Font(italic=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 16
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 14
    
    wb.save(output_filename)
    print(f"  Saved: {output_filename}")


def create_table2(results_all, results_pandemic, all_years_str, pandemic_years_str, output_filename):
    """
    Create Table 2: ALL, LT65 and GE65 with All Years and Pandemic Years side by side.
    
    Columns: Disease, W0, W1, W2, W2A (All Years), W0, W1, W2, W2A (Pandemic Years)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Disease_Weights_by_Age"
    
    row_num = 1
    
    for age_group in ['ALL', 'LT65', 'GE65']:
        if age_group not in results_all:
            continue
        
        rUDS_all = results_all[age_group]['rUDS']
        rADS_all = results_all[age_group]['rADS']
        
        rUDS_pan = results_pandemic[age_group]['rUDS'] if age_group in results_pandemic else None
        rADS_pan = results_pandemic[age_group]['rADS'] if age_group in results_pandemic else None
        
        # Section header
        age_label = {
            'ALL': 'Death counts for all ages, in thousands (% of W0)',
            'LT65': 'Death counts for under 65 years old, in thousands (% of W0)',
            'GE65': 'Death counts for 65 years old or older, in thousands (% of W0)'
        }[age_group]
        
        ws.cell(row=row_num, column=1, value=age_label)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
        row_num += 1
        
        # Sub-headers: All Years | Pandemic Years
        ws.cell(row=row_num, column=2, value=f"All Years ({all_years_str})")
        ws.cell(row=row_num, column=2).font = HEADER_FONT
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        
        ws.cell(row=row_num, column=6, value=f"Pandemic Years ({pandemic_years_str})")
        ws.cell(row=row_num, column=6).font = HEADER_FONT
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=row_num, start_column=6, end_row=row_num, end_column=9)
        row_num += 1
        
        # Column headers
        headers = ['Disease', 'W0', 'W1', 'W2', 'W2A', 'W0', 'W1', 'W2', 'W2A']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        row_num += 1
        
        # Data rows
        for letter in DISEASE_ORDER:
            # All years data
            w0_all = rUDS_all[letter]['First']
            w1_all = rUDS_all[letter]['W1']
            w2_all = rUDS_all[letter]['W2A']
            w2a_all = rADS_all[letter]['W2A']
            
            # Pandemic years data
            if rUDS_pan and rADS_pan:
                w0_pan = rUDS_pan[letter]['First']
                w1_pan = rUDS_pan[letter]['W1']
                w2_pan = rUDS_pan[letter]['W2A']
                w2a_pan = rADS_pan[letter]['W2A']
            else:
                w0_pan = w1_pan = w2_pan = w2a_pan = 0
            
            disease_name = DISEASE_NAMES.get(letter, letter)
            
            row_data = [
                disease_name,
                format_w0(w0_all),
                format_value(w1_all, w0_all),
                format_value(w2_all, w0_all),
                format_value(w2a_all, w0_all),
                format_w0(w0_pan),
                format_value(w1_pan, w0_pan) if w0_pan > 0 else "0 (+0)",
                format_value(w2_pan, w0_pan) if w0_pan > 0 else "0 (+0)",
                format_value(w2a_pan, w0_pan) if w0_pan > 0 else "0 (+0)"
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = THIN_BORDER
                if col == 1:
                    cell.font = Font(bold=True)
                else:
                    cell.alignment = Alignment(horizontal='right')
            
            row_num += 1
        
        row_num += 1  # Blank row between age groups
    
    # Footnote
    ws.cell(row=row_num, column=1, value="*endocrine nutritional and metabolic")
    ws.cell(row=row_num, column=1).font = Font(italic=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 16
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 14
    
    wb.save(output_filename)
    print(f"  Saved: {output_filename}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    args = parse_arguments()
    
    print("=" * 70)
    print("DISEASE WEIGHT TABLES GENERATOR")
    print("=" * 70)
    
    # Parse year ranges
    all_start, all_end = parse_year_range(args.all_years)
    pan_start, pan_end = parse_year_range(args.pandemic_years)
    
    all_years_str = args.all_years
    pandemic_years_str = args.pandemic_years
    
    # Read data
    print(f"\nReading: {args.filename}")
    df = pd.read_csv(args.filename, keep_default_na=False, na_values=[''])
    print(f"  Loaded {len(df):,} records")
    
    # Check required columns
    required = ['Count', 'rUDS', 'rADS', 'year', 'age_group']
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"ERROR: Missing columns: {missing}")
        print(f"  Available columns: {list(df.columns)}")
        sys.exit(1)
    
    total_deaths = df['Count'].sum()
    print(f"  Total deaths in file: {total_deaths:,}")
    
    # Run analyses
    print(f"\nAnalyzing All Years ({all_years_str})...")
    results_all = run_analysis(df, all_start, all_end)
    
    print(f"\nAnalyzing Pandemic Years ({pandemic_years_str})...")
    results_pandemic = run_analysis(df, pan_start, pan_end)
    
    # Output filenames
    if args.output:
        output_prefix = args.output
    else:
        output_prefix = "Disease_Weights_Tables"
    
    table1_file = f"{output_prefix}_Table1_All_Ages.xlsx"
    table2_file = f"{output_prefix}_Table2_by_Age_Period.xlsx"
    
    # Create tables
    print("\n" + "=" * 70)
    print("CREATING EXCEL TABLES")
    print("=" * 70)
    
    print("\nTable 1: ALL ages with All Years and Pandemic Years side by side")
    create_table1(results_all, results_pandemic, all_years_str, pandemic_years_str, table1_file)
    
    print("\nTable 2: LT65 and GE65 with All Years and Pandemic Years side by side")
    create_table2(results_all, results_pandemic, all_years_str, pandemic_years_str, table2_file)
    
    print("\n" + "=" * 70)
    print("DONE")
    print("=" * 70)
    print(f"\nOutput files:")
    print(f"  - {table1_file}")
    print(f"  - {table2_file}")


if __name__ == "__main__":
    main()
